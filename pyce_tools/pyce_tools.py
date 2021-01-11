import pandas as pd
import math
import plotly.express as px
import plotly.graph_objs as go
from plotly.subplots import make_subplots
import os
import numpy
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import scipy.stats as stats

class inp(object):
    
    var_names_uway = {
    'TRIPLET_TripletBeta660' : 'Triplet Beta',
    'TRIPLET_TripletCDOM' : 'CDOM',
    'TRIPLET_TripletChl' : 'Chl-a',
    'prokaryoticpico-syne' : 'Prok. Pico. Syn.',
    'nanophyto2-20um' : 'Nanophytoplankton',
    'picophyto<2um' : 'Picophytoplankton',
    'chla_cdom': 'chla/cdom',
    'SB21_SB21sal':'salinity'
    }

    
    def __init__(self, inp_type, inp_location, cyto_location, cyto_data, uway_bio_data, inp_data):
        '''
        Description
        ------------
        INP object. Will consist of INP of a specific type and location and bio data collected from a specific location.
        For example, an object named inp_uway would have inp of type seawater from location uway and cyto and uway bio data from uway. If you wanted to compare this
        INP data with cyto from a different location, a new INP object would need to be made.        
        
        Parameters
        ------------
        self : obj
            The INP object.
        inp_type : str
            The sample type for which this sample was collected. [seawater, aerosol]
        inp_location : str
            Where sample was collected. [uway, ASIT, wkbtsml, wkbtssw, bubbler, coriolis]
        cyto_location : 
            String to index the cyto_data df.
        cyto_data : df
            A dataframe of cyto data. Each row is an observation.
            Index should be named 'datetime' and columns are whatever measurements were taken. A location column must be present as it is indexed.
        uway_bio_data : df
            A dataframe of bio data. Each row is an observation.
            Index should be named 'datetime' and columns are whatever measurements were taken.
        inp_data : df
            A dataframe of INP data. Each row is an observation.
            Index should be named 'datetime'. Columns should include process, filtered, location, type, temp, concentration (inp/ml or inp/l) and uncertainty.
        '''
        self.inp_type = inp_type
        self.inp_location = inp_location
        self.uway_bio = uway_bio_data
        self.cyto_bio = cyto_data[cyto_data['location']==cyto_location]
        self.inp = inp_data[(inp_data['location']==inp_location)&(inp_data['type']==inp_type)]
        self.results = {}
    
    def corr(self, data, regime, temp):
        stat = pd.DataFrame()
        n = pd.DataFrame()
        pInput = pd.DataFrame()
        statCombined = pd.DataFrame()
        nCombined = pd.DataFrame()
        regime = str(regime)
        df=data
        df = df.select_dtypes(exclude=['object'])
        for column in df:
            if column == regime:
                continue
            try:
                stat = pd.DataFrame()
                pInput = df.loc[:,[regime,column]].dropna()
                ptest = stats.pearsonr(pInput[regime],pInput[column])
                stat[column]=ptest
                n = pInput.shape[0]
                stat=stat.append({column:n}, ignore_index=True)
                statT = stat.T
                statCombined = statCombined.append(statT)
            except ValueError:
                continue
                
        statCombined['variable'] = statCombined.index
        # Remove self-correlations
        statCombined['variable']=statCombined['variable'].astype(str)
        statCombined =statCombined[~statCombined['variable'].str.startswith('-')]
        # Calculate R^2 and name columns
        statCombined = statCombined.rename(columns={0:'R',1:'p', 2:'n'})
        statCombined['R^2'] = statCombined['R']**2
        
        # Add information
        statCombined['inp_temp']=temp
        
        return(statCombined, df)

    def sa_normalize(self, dA_total):
        '''Function takes a dA_total and inp and calculates surface area normalized in units of INP/cm2

        INPUT
        
        dA_total: total surface area integrated across all particle sizes for each day. Index should be a day number and SA should be in units of um2/cm3. If no index name, the function names for us.
        inp: INP/l of air. Columns should only be temperatures. Any other identifying information should be in the index. Index needs to include day. '''

        # ensure inp type is correct
        if self.inp_type != 'aerosol':
            return print('This function only works for aerosol type INP.')

        # rename dA_total index if needed
        if dA_total.index.name != 'day':
            dA_total.index.rename('day', inplace=True)
        
        # add a column of day
        if 'day' not in self.inp.columns:
            self.inp['day'] = self.inp.index.day.astype(str)

        # units of um2/cm3 convert to cm2/m3
        dA_total_prep = dA_total*1e-8*1e6

        # merge and set index
        self.inp = pd.merge(self.inp.reset_index(), dA_total_prep.reset_index(), on='day', how='outer')

        # calculate
        self.inp['inp_sa_normalized'] = self.inp['inp/m^3'].div(self.inp['SA'])

        return print('done')
        
    def correlations(self, temps, process, inp_units, dfs=None, size=None):
        
        self.results.update({process:{}})

        for temp in temps:
            self.results[process].update({temp:{}})
            
                
            # sort index
            self.inp = self.inp.sort_index()
            self.uway_bio = self.uway_bio.sort_index()
            self.cyto_bio = self.cyto_bio.sort_index()

             # merge INP and uway bio dataframes on date. This process is slightly different if the INP type is aerosol
            if self.inp_type =='aerosol':
                inp_uway_bio = pd.merge_asof(self.inp[(self.inp['temp']==temp)&(self.inp['process']==process)&(self.inp['size']==size)][inp_units], self.uway_bio, left_index=True, right_index=True, direction='nearest')
            
            else:
                inp_uway_bio = pd.merge_asof(self.inp[(self.inp['temp']==temp)&(self.inp['process']==process)][inp_units], self.uway_bio, left_index=True, right_index=True, direction='nearest')

            # merge INP with cyto dataframes on date
            inp_uway_bio_cyto_bio = pd.merge_asof(inp_uway_bio, self.cyto_bio, left_index=True, right_index=True, direction='nearest')
            
            df_corr = inp_uway_bio_cyto_bio

            # if any dfs were given, merge them here
            if dfs:
                for df in dfs:
                    df = df.sort_index()
                    df_corr = pd.merge_asof(df_corr, df, left_index=True, right_index=True, direction='nearest')
            
            [corrs, data_combined] = self.corr(df_corr, inp_units, temp)
            self.results[process][temp].update({'corrs':corrs})
            self.results[process][temp].update({'data':data_combined})

            
            print(f'Calculating correlations {process} INP samples of type={self.inp_type} at {temp}...Done!')
    
    def plot_corr_scatter(self, temp, units, processes, row_num):
        colors = ['steelblue','firebrick']
        tick_loc='inside'
        tick_width=2
        tick_length=10
        line_color='black'
        line_width=2

        xticks=10
        x_tick_angle=45
        x_tick_font_size=16
        x_tick_font_color='black'

        x_title_size=20
        x_title_color='black'

        yticks=10
        y_tick_angle = 90
        y_tick_font_size=16
        y_tick_font_color = "black"

        y_title_size=20
        y_title_color="black"
        
        vert_space = .10
        horz_space = .095

            
        # rows_needed = round(len(self.var_names_uway.keys())/3) + 1
        rows_needed = row_num

        fig = make_subplots(rows = rows_needed, cols = 3, shared_xaxes=False, shared_yaxes=False,
                            horizontal_spacing = horz_space, vertical_spacing=vert_space, print_grid=False
                            )

        row_num = 1
        col_num = 1

        for variable in self.var_names_uway:
            
            ytitle=f'INP<sub>{self.inp_location}</sub>'
            xtitle = variable
            
            i = 0

            for temperature in temp:

                for process in processes:
                    datay_string = temperature
                    datax = self.results[process][temperature]['data'].dropna(subset=[units, variable])


                    fig.add_trace(go.Scatter(
                            name = process, 
                            y= datax[units],
                            x= datax[variable],
                            mode="markers", connectgaps=True,
                            marker=dict(color=colors[i], symbol='circle-open')),
                            row=row_num,
                            col = col_num)


                    data_trendline=self.results[process][temperature]['data'].dropna(subset=[variable, units])

                    fig_trend = px.scatter(data_trendline, x=variable, y=units, trendline="ols")
                    trendline = fig_trend.data[1] 
                    fig_trend.update_traces(line=dict(color=colors[i],dash='dash'))
                    fig.add_trace(go.Scatter(trendline), row=row_num, col=col_num)



                    fig.update_xaxes(row=row_num, col = col_num, 
                            title=dict(text=xtitle, font=dict(size=x_title_size, color=x_title_color)), 
                            exponentformat='power', tickangle = x_tick_angle,
                            ticks=tick_loc, nticks=xticks, tickwidth=tick_width, ticklen=tick_length, showline=True, 
                            linecolor=line_color,linewidth=line_width,  
                            tickfont=dict(size=x_tick_font_size, color=x_tick_font_color))

                    fig.update_yaxes(row=row_num, col = col_num, 
                            title=dict(text=ytitle,font=dict(size=y_title_size, color=y_title_color)),
                            exponentformat='power',  
                            ticks=tick_loc, nticks=10, tickwidth=tick_width, ticklen=tick_length,
                            tickfont=dict(size=y_tick_font_size, color=y_tick_font_color),
                            showline=True, linecolor=line_color, linewidth=line_width)

                    fig.update_traces(marker=dict(size=10,
                                line = dict(width=2, color='DarkSlateGrey')))


                # make for all situations. only bold if significant.
                    if self.results[process][temperature]['corrs'].loc[variable,'p'] < .05:
                        anno_text = '<b>R<sup>2</sup>=' + str(round(self.results[process][temperature]['corrs'].loc[variable,'R^2'],2))

                    elif self.results[process][temperature]['corrs'].loc[variable,'p'] > .05:
                        anno_text = 'R<sup>2</sup>=' + str(round(self.results[process][temperature]['corrs'].loc[variable,'R^2'],2))
                    
                    x_val = [0.01, 0.41, 0.80]
                    y_val = round((1-(row_num-1)*(1/(rows_needed-.35)))-i*.035,3)


                    fig.add_annotation(
                        showarrow=False,
                        xref="paper",
                        yref="paper",
                        x=x_val[col_num-1],
                        y=y_val,
                        text = anno_text,
                        font=dict(size=18, color=colors[i]))


                    fig.update_layout(width=1500, height=400*rows_needed, template='plotly_white',showlegend=False, title=f'Correlation Scatter Plots for {self.inp_location} at {temperature}')
                    i+=1
                            

            col_num+=1
            if col_num%4 == 0:
                col_num=1
                row_num +=1
        #fig.write_image(selection+'.png', scale=3)
        #fig.show()
        return fig

    def plot_ins_inp(self):
        if self.inp_type != 'aerosol':
            return print('INP object must be aerosol type. To plot seawater type INP, use plot_sml_inp or plot_ssw_inp.')
        fig=go.Figure()

        fig.add_trace(
            go.Scatter(name='Unheated',
                x=self.inp[self.inp['process'] == 'UH']['temp'],
                y=self.inp[self.inp['process'] == 'UH']['inp_sa_normalized'],
                mode='markers',
                marker=dict(size=7, color='blue',line = dict(width=1, color='black'))))

        fig.add_trace(
            go.Scatter(name='Heated',
                x=self.inp[self.inp['process'] == 'H']['temp'],
                y=self.inp[self.inp['process'] == 'H']['inp_sa_normalized'],
                mode='markers',
                marker=dict(size=7, color='red',line = dict(width=1, color='black'))))

        fig.add_trace(
            go.Scatter(name='DeMott et al. (2016) Caribbean, Arctic, Pacific, Bering Sea - Ambient',
                x=[-13, -25, -15, -13],
                y=[1, 900, 20, 1],
                mode="lines", connectgaps=True, line=dict(color="tomato", width=1.5))
        )

        fig.add_trace(
            go.Scatter(name='DeMott et al. (2016) SIO Pier seawater - Waveflume/MART',
                x=[-15, -25, -23, -13, -15],
                y=[.2, 40, 100, 1, .2],
                mode="lines", connectgaps=True)
        )

        fig.add_trace(
            go.Scatter(name='McCluskey et al. (2017) SIO Pier seawater - Waveflume',
                x=[-10, -23, -30, -30, -10, -10],
                y=[1.5, 20, 300, 500, 70, 1.5],
                mode="lines", connectgaps=True, line=dict(color="deepskyblue", width=1.5))
        )

        fig.add_trace(
            go.Scatter(name='McCluskey et al. (2018b) Southern Ocean - Ambient',
                x=[-12, -18, -26, -26, -12],
                y=[.4, 0.2, 100, 8000, 0.4],
                mode="lines", connectgaps=True, line=dict(color="darkgoldenrod", width=1.5))
        )

        fig.add_trace(
            go.Scatter(name='Gong et al. (2020) Cabo Verde - Ambient',
                x=[-10, -11, -20, -19, -10],
                y=[2, .3, 12, 100, 2],
                mode="lines", connectgaps=True, line=dict(color="seagreen", width=1.5))
        )

        fig.update_yaxes(type='log', title='INP per cm<sup>2</sup> of SSA Surface (D<sub>p</sub> = 10-500nm)', exponentformat='power')
        fig.update_xaxes( title='Temperature (\u00B0C)',range=[-30.5,-5], exponentformat='power')

        fig.update_yaxes(linewidth=2, tickwidth=2, ticklen=10, ticks='inside',
                        showline=True,linecolor='black', tickfont=dict(size=16), title_font=dict(size=16))

        fig.update_xaxes(linewidth=2, tickwidth=2, ticklen=10,
                        ticks='inside', showline=True, linecolor='black', tickfont=dict(size=16), title_font=dict(size=16))
        
        fig.update_layout(template='plotly_white', 
            width=700, height=600, 
            showlegend=True, legend=dict(title='', font=dict(size=10, color='black'), x=0.2, y=1.2, bordercolor="black", borderwidth=1))

        return fig

def calculate_raw_blank(type_, process, location, sample_name, collection_date, analysis_date, issues, num_tubes, vol_tube = 0.2, rinse_vol = 20, size = None):
    '''
    Description
    ------------
    Loads raw data from LINDA BLANK experiments and creates a 'calculated' INP data file using given arguments.
    Saves the output as an XLSX file which can be later used as the blank in sample calculations of LINDA experiments.

    Paths
    ------------
    raw input data: \\[PROJECT_ROOT]\\data\\raw\\IN\\blank\\[FILE]
    calculated output file: \\[PROJECT_ROOT]\\data\\interim\\IN\\calculated\\blank\\[FILE]
    
    Recent Updates
    -------------
    02/11/2020 - Improved documentation.
    
    Parameters
    ------------
    type_ : str
        The sample type for which this blank was collected. Note that mq is for any seawater samples, whether they be from the underway or workboat. [aerosol, mq] 
    process : str
        Identify whether the sample has been left unfiltered (uf) or filtered (f). Unheated (UH) and heated (H) processes are already included in the file. [uf,f]
    location : str
        [bubbler, coriolis, mq, mq_wboat]
    sample_name : str
        Sample source name as seen on the vial. There is no rule on this, as it is simply stored as metadata, but ideally the sample name is following some kind of consistent standard.
    collection_date : str
        Sample collection date. Make note of your timezone, as the code does not assume one. 
        This is used to help locate the raw data file. [DDMMYYYY HHhMM for seawater samples dayXX for aerosols].
    analysis_date : str
        Date of LINDA analysis in NZST time. [DD/MM/YY]
    issues : str
        Issues noted in the LINDA analysis log. [DEFAULT = None]
    num_tubes : int
        Number of tubes per heated/unheated analysis. [DEFAULT = 26]
    vol_tube : int
        Volume in ml of sample solution per tube. [DEFAULT = 0.2]
    rinse_vol : int
        Volume in ml of mq water used for rinsing filters, if the sample type_ makes use of a filter.
    size : str
        Size of particles for filter samples if sample was size resolved. [super, sub]
    
    Returns
    ------------
    xlsx
        A spreadsheet of calculated blank data.
    '''
    

    # load raw data depending on sample type
    if type_ == 'mq':
        date = collection_date[:6]
        inpath = '..\\data\\raw\\IN\\blank\\' + location + '_'+ 'blank' + '_' + process + '_' + date + '.csv'
        template = pd.read_excel('..\\in_calculation_template.xlsx', skiprows=1)
    
    if type_ == 'aerosol':
        date = collection_date
        inpath = '..\\data\\raw\\IN\\blank\\' + location + '_'+ 'blank' + '_' + process + '_' + size + '_' + date + '.csv'
        template = pd.read_excel('..\\in_calculation_template_aerosols.xlsx', skiprows=1)
    
    raw = pd.read_csv(inpath, sep = ' ', header = None, parse_dates=False)
    
    # create a datetime df to split up the date and time into separate columns, then insert them into the raw spreadsheet
    datetime_col = raw[0].str.split(' ', expand=True)
    raw.insert(0, 'day',datetime_col[0])
    raw[0]=datetime_col[1]
    
    # drop the random extra column at the end
    raw=raw.drop(columns=61)
    
    # set the column names so they match those in the template
    raw.columns = template.columns
    
    # create metadata dict depending on sample type
    if type_ == 'mq_wboat' or type_ == 'mq':
            meta_dict = {
                'raw data source': inpath,
                'type':type_,
                'location':location,
                'process':process,
                'sample source name': sample_name,
                'sample collection date': collection_date,
                'sample analysis date': analysis_date,
                '# tubes': num_tubes,
                'ml/tube': vol_tube,
            }
    
    if type_ == 'aerosol':
        meta_dict = {
            'raw data source': inpath,
            'type':type_,
            'location':location,
            'process':process,
            'sample source name': sample_name,
            'sample collection date': collection_date,
            'sample analysis date': analysis_date,
            '# tubes': num_tubes,
            'ml/tube': vol_tube,
            'rinse volume': rinse_vol,
            'size': size,
        }
    
    
    # insert the raw data into the template 
    if type_ == 'mq_wboat' or type_ == 'mq':
        template = load_workbook('..\\in_calculation_template.xlsx')
    if type_ == 'aerosol':
        template = load_workbook('..\\in_calculation_template_aerosols.xlsx')
    template.remove(template["data.csv"])
    sheet = template.create_sheet('data.csv')
    for row in dataframe_to_rows(raw, index=False, header=True):
        sheet.append(row)
    sheet.insert_rows(idx=0)
    
    # add metadata to spreadsheet
    if type_ == 'aerosol':
        row = 0
        for key, value in meta_dict.items():
            template['summary_UF_UH']['v'][row].value = key
            template['summary_UF_UH']['w'][row].value = value
            template['summary_UF_H']['v'][row].value = key
            template['summary_UF_H']['w'][row].value = value
            row +=1
    if type_ =='mq_wboat' or type_ =='mq':
        row = 0
        for key, value in meta_dict.items():
            template['summary_UF_UH']['z'][row].value = key
            template['summary_UF_UH']['aa'][row].value = value
            template['summary_UF_H']['z'][row].value = key
            template['summary_UF_H']['aa'][row].value = value
            row +=1
    # save calculated report file to the appropriate folder
    if type_ == 'mq_wboat' or type_ == 'mq':
        template.save('..\\data\\interim\\IN\\calculated\\blank\\'+type_ + '_'+'blank'+'_' + process + '_' + date+'_calculated.xlsx')
    if type_ == 'aerosol':
        template.save('..\\data\\interim\\IN\\calculated\\blank\\'+location + '_'+'blank'+'_' + process + '_' + size + '_'+ date + '_calculated.xlsx')
    
    return print('...Raw blank data calculated!')

def calculate_raw(blank_source, type_, location, process, sample_name, 
                  collection_date, analysis_date, issues, num_tubes, vol_tube = 0.2, rinse_vol = 20, size = None,
                  flow_start = None, flow_stop = None, sample_stop_time = None):
    '''
    Description
    ------------
    Creates an XLSX spreadsheet of blank corrected, calculated INP data for samples using given args. 
    Resulting spreadsheet has a seperate tab for unheated and heated samples, with respective metadata in each.
    Saves the output to interim calculated folder --> data/interim/IN/calculated/[seawater or aerosols].
    
    Paths
    ------------
    raw input data: \\[PROJECT_ROOT]\\data\\raw\\IN\\[SAMPLE_TYPE]\\[FILE]
    calculated output file: \\[PROJECT_ROOT]\\data\\interim\\IN\\calculated\\[SAMPLE_TYPE]\\[SAMPLE_LOCATION]\\[FILE]
    
    TODO
    ------------
    Take average of all blanks and insert rather than just using one blank file right now.
    Include functionality for filtered seawater samples.
    
    Recent Updates
    -------------
    02/11/2020 - Further improved documentation.

    15/09/2020 - Improved documentation.
    
    11/09/2020 - Added code to create uncertainty for SEAWATER and AEROSOL samples based on Wilson Score. Score is calculated within the template notebook. Fixed a few bugs.

    21/08/2020 - Added code on seawater and aerosol outpath to account for location (i.e., wboat or uway). Now saves to 
    a separate folder within the respective type folder.

    11/08/2020 - Added functionality for Coriolis samples.
    
    23/07/2020 - Added functionality for start and stop time of aerosol sampling. Now automatically calculates sampling time in
    minutes based on input args. 
    
    Parameters
    ------------
        blank_source : str
            Path to source of calculated blank data. Currently accepts one file, but need to account for average of multiple files. 
            (e.g., '..\\data\\interim\\IN\\calculated\\blank\\mq_blank_uf_120620_calculated.xlsx')
        type_ : str
            The sample type for which this blank was collected. [seawater, aerosol]
        location : str
            Where sample was collected. [uway, ASIT, wboatsml, wboatssw, bubbler, coriolis]
        process : str
             Identify whether the sample has been unfiltered or filtered. Unheated and heated processes are already included in the file. [uf,f]
        sample_name : str 
            sample source name as seen on the vial. 
        collection_date : str 
            sample collection date in NZST. [DDMMYYYY HHhMM for seawater and aerosol samples.]
        analysis_date : str
            LINDA analysis date in NZST. [DD/MM/YY]
        issues : str
            Issues noted in the LINDA analysis log. [DEFAULT = None]
        num_tubes : int
             Number of tubes per heated/unheated analysis. [DEFAULT = 26]
        vol_tube : int
            Volume in ml of sample solution per tube. [DEFAULT = 0.2]
        rinse_vol : int
            Volume in ml of mq water used for rinsing filters (aerosol sample types only).
        size : str
            Size of particles for filter samples (if applicable). Defaults to None if not given. Only used for aerosol samples. [super, sub]
        flow_start : float 
            Flow rate in LPM at start of sampling. Only used for aerosol samples.
        flow_stop : float
            Flow rate in LPM at end of sampling. Only used for aerosol samples.
        sample_stop_time : str
            Time in NZST at which sample collection was halted. Only valid for aerosol collections. [DDMMYYYY HHhMM]
    '''

    # Dictionary for mapping actual datetime of sample to the variable collection_date. Used for locating files (due to my poor file naming scheme) and can be ignored by anyone not working with
    # coriolis samples from the 2020 S2C Tangaroa Cruise.
    coriolis_day_date = {
        '18032020 13h00':'tg_04',
        '21032020 13h00':'tg_6',
        '25032020 14h00':'tg_9-2',
        '22032020 14h00':'tg_7',
        '23032020 13h00':'tg_8-2',
        '19032020 15h00':'tg_5-7',
    }
    
    
    # load raw data depending on sample type (seawater vs aerosol)
    if type_ == 'seawater':
        # extract date and time from input parameters
        date = collection_date[:6]
        time = collection_date[9:11]+collection_date[12:]
        # use input parameters to build path to source file
        inpath = '..\\data\\raw\\IN\\' + type_ + '\\' + type_ + '_' + location + '_' + process + '_' + date + '_' + time + '.csv'
        # load analysis template
        template = pd.read_excel('..\\in_calculation_template.xlsx', skiprows=1)
    
    if type_ == 'aerosol':
        if location == 'bubbler':
            # extract date and time from input parameters
            date = collection_date[:6]
            time = collection_date[9:11]+collection_date[12:]
            # use input parameters to build path to source file
            inpath = '..\\data\\raw\\IN\\' + type_ + '\\' + type_ + '_' + location + '_' + process + '_' + size + '_'+ date + '_' + time + '.csv'
            # load analysis template
            template = pd.read_excel('..\\in_calculation_template_aerosols.xlsx', skiprows=1)
        if location == 'coriolis':
            # extract date and time from input parameters
            date = coriolis_day_date[collection_date]
            # use input parameters to build path to source file
            inpath = '..\\data\\raw\\IN\\' + type_ + '\\' + type_ + '_' + location + '_' + process + '_' + date + '.csv'
            # load analysis template
            template = pd.read_excel('..\\in_calculation_template_aerosols.xlsx', skiprows=1)
    
    # read in the raw data csv
    raw = pd.read_csv(inpath, sep = ' ', header = None, parse_dates=False)
    

    # create a datetime df to split up the date and time into separate columns, then insert them
    datetime_col = raw[0].str.split(' ', expand=True)
    raw.insert(0, 'day',datetime_col[0])
    raw[0]=datetime_col[1]

    # drop the extra empty column at the end of every data file
    raw=raw.drop(columns=61)
    
    # set the column names so they match those in the template
    raw.columns = template.columns

    # create metadata dict depending on sample type
    if type_== 'seawater':
            meta_dict = {
                'raw data source': inpath,
                'type':type_,
                'location':location,
                'process':process,
                'sample source name': sample_name,
                'sample collection date': collection_date,
                'sample analysis date': analysis_date,
                '# tubes': num_tubes,
                'ml/tube': vol_tube,
                'issues':issues,
                'sigma' : 1.96
            }

    if type_ == 'aerosol':
        # Calculate sampling time in minutes
        t_start=datetime.datetime.strptime(collection_date, '%d%m%Y %Hh%M')
        t_end=datetime.datetime.strptime(sample_stop_time, '%d%m%Y %Hh%M')
        sample_time = t_end - t_start 
        sample_time_mins = sample_time.total_seconds()/60
        
        # Create metadata dictionary
        meta_dict = {
            'raw data source': inpath,
            'type':type_,
            'location':location,
            'process':process,
            'sample source name': sample_name,
            'sample collection date': collection_date +  'through ' + sample_stop_time,
            'sample analysis date': analysis_date,
            '# tubes': num_tubes,
            'ml/tube': vol_tube,
            'rinse volume': rinse_vol,
            'size': size,
            'avg_flow': (flow_start+flow_stop)/2,
            'sample collection time (minutes)': sample_time_mins
        }
        meta_dict['total air volume'] = meta_dict['avg_flow'] * meta_dict['sample collection time (minutes)']
        meta_dict['issues'] = issues
        meta_dict['sigma'] = 1.96

    
    # insert the raw data into the template 
    if type_ == 'seawater':
        template = load_workbook('..\\in_calculation_template.xlsx')
    if type_ == 'aerosol':
        template = load_workbook('..\\in_calculation_template_aerosols.xlsx')
    template.remove(template["data.csv"])
    sheet = template.create_sheet('data.csv')
    for row in dataframe_to_rows(raw, index=False, header=True):
        sheet.append(row)
    sheet.insert_rows(idx=0)
    
    
    # add metadata to spreadsheet - one in each of the two process sheets (UF_UH and UF_H)
    if type_ == 'seawater':
        row = 0
        for key, value in meta_dict.items():
            template['summary_UF_UH']['z'][row].value = key
            template['summary_UF_UH']['aa'][row].value = value
            template['summary_UF_H']['z'][row].value = key
            template['summary_UF_H']['aa'][row].value = value
            row += 1
    if type_ == 'aerosol':
        row = 0
        for key, value in meta_dict.items():
            template['summary_UF_UH']['v'][row].value = key
            template['summary_UF_UH']['w'][row].value = value
            template['summary_UF_H']['v'][row].value = key
            template['summary_UF_H']['w'][row].value = value
            row += 1
    
    
    # read and add blank data.
    blank_uf_uh = pd.read_excel(blank_source, sheet_name='summary_UF_UH')
    blank_uf_h = pd.read_excel(blank_source, sheet_name='summary_UF_H')
    row = 1
    for x in blank_uf_uh['N(frozen)']:
        template['summary_UF_UH']['f'][row].value = x
        row += 1
    row = 1
    for x in blank_uf_h['N(frozen)']:
        template['summary_UF_H']['f'][row].value = x
        row += 1
    
    
    # Save output depending on IN type
    if type_ == 'seawater':
        template.save('..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'+type_ + '_' + location + '_' + process + '_' + date + '_' + time +'_calculated.xlsx')
        save_path = '..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'+type_ + '_' + location + '_' + process + '_' + date + '_' + time +'_calculated.xlsx'
    if type_ == 'aerosol':
        if location == 'bubbler':
            template.save('..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'+type_ + '_' + location + '_' + process + '_' + size + '_' + date + '_' + time + '_calculated.xlsx')
            save_path = '..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'+type_ + '_' + location + '_' + process + '_' + size + '_' + date + '_' + time + '_calculated.xlsx'
        if location == 'coriolis':
            template.save('..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'+type_ + '_' + location + '_' + process  + '_' + date+'_calculated.xlsx')
            save_path = '..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'+type_ + '_' + location + '_' + process  + '_' + date+'_calculated.xlsx'
    

    return print(f'...IN data calculated!\nCalculated report file saved to {save_path}.')

def clean_calculated_in(type_, location):
    '''
    Description
    ------------
    Creates an XLSX spreadsheet of cleaned data ready for analysis. 
    
    Paths
    ------------
    raw input data: \\[PROJECT_ROOT]\\data\\interim\\IN\\calculated\\[SAMPLE_TYPE]\\[SAMPLE_LOCATION]\\[FILE]
    cleaned output file: \\[PROJECT_ROOT]\\data\\interim\\IN\\cleaned\\combinedtimeseries\\[SAMPLE_TYPE]\\[FILE]
    
    TODO
    ------------
    aerosol coriolis sample files should be changed to have similar information and format as seawater sample files.

    Recent Updates
    -------------
    09/11/2020 - Fixed aerosol bubbler process to have similar information and format as seawater sample files.
    02/11/2020 - Added documentation.
    
    Parameters
    ------------
        type_ : str
            The sample type for which this blank was collected. [seawater, aerosol]
        location : str
            Where sample was collected. [uway, ASIT, wkbtsml, wkbtssw, bubbler, coriolis]
    '''
    
    
    # create a dataframe that will contain all of the data from each separate calculated file.
    big_df = pd.DataFrame()
    
    # cycle through all calculated report files in the folder
    for file in os.listdir('..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'):
        
        # account for unheated and heated processes
        procs = ['UH','H']

        if type_=='seawater':
            
            for process in procs:
                
                # load the file
                df = pd.read_excel('..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'+file, sheet_name='summary_UF_'+process)
                
                # rename annoying column names
                df['T (*C)'] =df['T (*C)'].round(1)
                
                # create a df for transforming 
                current = pd.DataFrame()
                
                # create transposed df and give it appropriate column names
                fd = df.T
                fd.columns = fd.loc['T (*C)',:]
                
                # add data to dataframe
                current= current.append(fd.loc['IN/ml',:], ignore_index=True)
                current['datetime'] = fd.iloc[-1,4]
                current['time'] = current['datetime'].iloc[0][9:]
                
                # turn columns into strings so they aren't ints
                current.columns = current.columns.astype(str)
                
                # add label data
                current['process'] = process
                current['type'] = type_
                current['location'] = location
                current['filtered'] = 'uf'
                
                # append this to the final dataframe
                big_df=big_df.append(current)
        
        elif type_ == 'aerosol' and location == 'bubbler':
            
            for process in procs:
                
                # load the file
                df = pd.read_excel('..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'+file, sheet_name='summary_UF_'+process)
                
                # rename annoying column names
                df['T (*C)'] =df['T (*C)'].round(1)
                
                # create a df for transforming 
                current = pd.DataFrame()
                
                # create transposed df and give it appropriate column names
                fd = df.T
                fd.columns = fd.loc['T (*C)',:]
                
                # add data to current
                current=current.append(fd.loc['IN/L',:], ignore_index=True)
                current['datetime'] = fd.iloc[-1,4][0:14]
                current['start_date'] = fd.iloc[-1,4][0:14]
                current['stop_date'] = fd.iloc[-1,4][21:]

                # turn columns into strings so they aren't ints
                current.columns = current.columns.astype(str)

                # add label data
                if 'super' in file:
                    current['size'] = 'super'
                if 'sub' in file:
                    current['size'] = 'sub'
                current['process'] = process
                current['type'] = type_
                current['location'] = location
                current['filtered'] = 'uf'
                
                # append this to the final big_df
                big_df=big_df.append(current)
        
        elif type_=='aerosol' and location == 'coriolis':
            
            for process in procs:
                
                # load the file
                df = pd.read_excel('..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'+file, sheet_name='summary_UF_'+process)
                # rename annoying column names
                df['T (*C)'] =df['T (*C)'].round(1)
                # create a df for transforming 
                current = pd.DataFrame()
                # create transposed df and give it appropriate column names
                fd = df.T
                fd.columns = fd.loc['T (*C)',:]
                
                # add data to current
                current=current.append(fd.loc['IN/L (INP per liter of air)',:], ignore_index=True)
                current['date'] = fd.iloc[-1,4]
                current['hour'] = current['date'].iloc[0][9:]
                current['start_date'] = current.loc[0,'date'][0:14]
                current['stop_date'] = current.loc[0,'date'][23:]
                
                # add label data
                if 'super' in file:
                    current['size'] = 'super'
                if 'sub' in file:
                    current['size'] = 'sub'

                # turn columns into strings so they aren't ints
                current.columns = current.columns.astype(str)
                current['process'] = process
                
                # append this to the final big_df
                big_df=big_df.append(current)

    # save output to combined time series folder
    strt=big_df['datetime'].min()[0:8]
    end=big_df['datetime'].max()[0:8]
    out_name = strt+'_'+end
    big_df.to_csv('..\\data\\interim\\IN\\cleaned\\combinedtimeseries\\'+type_+'\\'+location+'_'+out_name+'.csv', index=False)
    #big_df.date=pd.to_datetime(big_df.datetime, dayfirst=True, format='%d%m%Y %Hh%M')

def clean_inverted(inpath, nbins, outpath):
    '''
    Description
    ------------
    Accepts inverted scanotron data files from a specified given folder. Appends them into one dataframe and 
    sends them out to /interim/scanotron/combinedtimeseries/ folder. Also returns the completed dataframe as a variable for immediate use, as well as the file name and dLogDp value.
    
    Paths
    ------------
    inverted scanotron input data folder path: ..\\data\\interim\\"+instr+"\\inverted\\pro\\BHS\\
    calculated output file: ..\\data\\interim\\'+instr+'\\combinedtimeseries\\BHS\\[FILE]

    Parameters
    ------------
    inpath : str
        Path to inverted scanotron data files. [example: '..\\data\\interim\\"+instr+"\\inverted\\pro\\BHS\\']
    nbins : int
        Number of diameter bins for scanotron.
    outpath : str
        Desired location for the combined time series csv file. [example: '..\\data\\interim\\'+instr+'\\combinedtimeseries\\BHS\\']

    Returns
    ------------
    dfBig
        Dataframe of combined time series of scanotron data. Rows are timestring and columns include time, diameters, and year, month, day,	hour, minute, second, pex, tex, rhsh, tgrad, nb, dbeg, dend, conctotal.
    outName
        Name of the file that is saved to the computer.
    dLogDp
        Integer of dLogDp.
    '''
    
    dfBig=pd.DataFrame()
    path=inpath

    # Read all the files in the folder defined by inpath variable.
    for file in os.listdir(path):
        if file.endswith('.csv'):
            df = pd.read_csv(path+file, skiprows=5,header=None,sep='\t')
            dfBig = dfBig.append(df)
            # Read in column names
            columns = pd.read_csv(path+file, skiprows=3,nrows=0,sep='\t').columns.tolist()
            # Remove all bad chars from column names
            for name in range(len(columns)):
                columns[name]=(columns[name]).strip()
                columns[name]=(columns[name]).replace("#","")
                columns[name]=columns[name].lower()
    # Count number of missing column names (these are due to the size bins of the data)
    num_missing_cols=nbins
    # Calculate and add in new column names based on the size of the bins
    start_bin = df.iloc[1,11]   # Smallest Dp
    end_bin = df.iloc[1,12]     # Largest Dp
    dLogDp=(math.log10(end_bin)-math.log10(start_bin))/(num_missing_cols-1)
    num = math.log10(start_bin)
    LogDp = [math.log10(start_bin)]
    while num < math.log10(end_bin):
        num = num + dLogDp
        LogDp.append(num)
    Dp = [10**j for j in LogDp]
    Dp=[round(x) for x in Dp]
    columns.remove('conc...')
    dfBig.columns = columns + Dp
    # Create datetimes
    dfBig=dfBig.rename(columns={'yr':'year','mo':'month','dy':'day','hr':'hour','mn':'minute','sc':'second'})
    dfBig['time']=pd.to_datetime(dfBig[['year', 'month', 'day', 'hour','minute','second']])
    dfBig['timeString'] = dfBig['time'].dt.strftime('%Y-%m-%d %H:%M:%S')
    # Set df index as datetime for easy parsing
    dfBig=dfBig.set_index(['timeString'])
    # Save as csv
    #dfBig.to_csv(outPath+'out.csv')
    strt=dfBig.index[0]
    end=dfBig.index[-1]
    outName = strt[0:10]+'_'+end[0:10]
    dfBig.to_csv(outpath+outName+'.csv')
    return dfBig, outName, dLogDp

def clean_magic(inpath, outpath, timezone):
    '''
    Description
    ------------

    Loads all raw magic CPC data files, cleans it up, and appends it into one file.
    Returns the cleaned dataset to chosen outpath as csv file. 
    
    The steps of the cleaning process are as follows:
        1) open all data files in the data/raw folder path
        2) append all data files into one df
        3) remove bad chars in column names
        4) create a timeString column in UTC time
        5) save df to csv in specified folder
    
    Parameters
    ------------
    inpath : str
         location where raw csv file is found.
    outpath : str
        location where cleaned csv file is saved.
    
    Returns
    ------------
    dfBig (df): the df that was just saved to a folder
    outName (str): string of the start and end datetimes
    
    '''
    dfBig=pd.DataFrame()
    path= inpath

    #Read in all the files in a given folder.
    for file in os.listdir(path):
        if file.endswith('.csv'):
            df = pd.read_csv(path+file, sep='\t', parse_dates=['# UTC               '], skiprows=3)
            dfBig = dfBig.append(df)
            # Read in column names
            columns = pd.read_csv(path+file,skiprows=3,nrows=0,sep='\t').columns.tolist()
            # Remove all bad chars from column names
            for name in range(len(columns)):
                columns[name]=(columns[name]).strip()
                columns[name]=(columns[name]).replace("#","")
                columns[name]=(columns[name]).replace(" ","")
                columns[name]=columns[name].lower()
                columns[name]=(columns[name]).replace("utc","time")
    dfBig.columns = columns
    dfBig.time=dfBig.time+DateOffset(hours=1)
    dfBig.set_index('time',inplace=True)
    dfBig=dfBig.tz_localize(None)
    dfBig=dfBig.tz_localize(timezone)
    dfBig.reset_index(inplace=True)
    dfBig['timeString'] = dfBig['time'].dt.strftime('%Y-%m-%d (%H:%M:%S)')
    dfBig=dfBig.set_index(['timeString'])
    strt=dfBig.index[0]
    end=dfBig.index[-1]
    outName = strt[0:10]+'_'+end[0:10]
    dfBig.to_csv(outpath+outName+'.csv')
    return dfBig, outName

def load_scano_data(date_string, instrument):
    '''
    Description
    ------------
    Loads interim scanotron data that has already been pre-processed using the clean_inverted function. Returns two dataframes of dN and dNdLogDp where rows are time and columns are diameters.
    
    Paths
    ------------
    raw input data: \\[PROJECT_ROOT]\\data\\interim\\scanotron\\combinedtimeseries\\[FILE]

    Parameters
    ------------
    date_string : str
        Dates of scanotron data that are requested. [YYYY-MM-DD_YYYY_MM_DD]
    instrument : str
        Instrument that is being loaded. [scanotron]

    Returns
    ------------
    dN
        A dataframe of particle counts where rows are dates and columns are diameters.
    dNdLogDp
        A dataframe of log-normalized particle counts where rows are dates and columns are diameters.
    '''
    # Read in the cleaned and combinedtimeseries scanotron data
    df=pd.read_csv(
        '..\\data\\interim\\'+instrument+'\\combinedtimeseries\\'+date_string+'.csv',
        parse_dates=['time'])
    # Create a dNdlogDp dataframe
    dNdLogDp_full  = df.set_index('time').loc[:,'10':]
    dLogDp=(math.log10(df.dend[1])-math.log10(df.dbeg[1]))/(df.nb[1]-1)
    # Create a dN dataframe
    dN=df.set_index('time').loc[:,'10':]*dLogDp
    return dN, dNdLogDp_full

def surface_area(smps_daily_mean_df, smps_daily_std_df, nbins):
    '''
    '''
     # Calculate log of particle diameter bin size (dLogDp)
    start_bin = smps_daily_mean_df.index[0]   # Smallest Dp
    end_bin = smps_daily_mean_df.index[-1]   # Largest Dp
    dLogDp=(math.log10(end_bin)-math.log10(start_bin))/(nbins-1)

    # Convert particle diameters to meters from nm
    smps_daily_mean_df.reset_index(inplace=True)
    smps_daily_mean_df['Dp (m)']=smps_daily_mean_df['Dp']*1e-9

    smps_daily_std_df.reset_index(inplace=True)
    smps_daily_std_df['Dp (m)'] = smps_daily_std_df['Dp']*1e-9

    # Create particle diameter squared (Dp2) in units of meters^2
    smps_daily_mean_df['Dp2 (m2)'] = smps_daily_mean_df['Dp (m)']**2
    smps_daily_std_df['Dp2 (m2)'] = smps_daily_std_df['Dp (m)']**2

    # Calculate factor A for particle surface area calculation in m^2, which is just diameter squared times Pi
    smps_daily_mean_df['factorA'] = smps_daily_mean_df['Dp2 (m2)']*3.14159
    smps_daily_mean_df.set_index('Dp',inplace=True)

    smps_daily_std_df['factorA'] = smps_daily_std_df['Dp2 (m2)']*3.14159
    smps_daily_std_df.set_index('Dp', inplace=True)

    # Calculate particles/cm^3 of air (un-normalized) by multiplying through by the normalizing constant (log of particle diameter bin size)
    dN=smps_daily_mean_df.iloc[:,:-3]*dLogDp
    dN_std = smps_daily_std_df.iloc[:,:-3]*dLogDp

    # Calculate surface area per unit volume of air (m^2/cm^3)
    dA=dN.mul(smps_daily_mean_df['factorA'],0)
    dA_std=dN_std.mul(smps_daily_std_df['factorA'],0)

    # Convert surface area units from m^2/cm^3 to um^2/cm^3
    dA=dA*1e12
    dA_std=dA_std*1e12

    # Normalize by log of particle diameter bin size (dLogDp) for surface area          distribution (nm^2/cm^3)
    dAdLogDp = (dA*1e6)/dLogDp
    dAdLogDp_std = (dA_std*1e6)/dLogDp

    # Sum all size bins to calculate total surface area (um^2/cm^3)
    dA_total=dA.sum(axis=0).to_frame()
    dA_total.rename(columns={0:'SA'},inplace=True)
    dN_total=dN.sum(axis=0).to_frame()
    dN_total.rename(columns={0:'DN'},inplace=True)
    return dAdLogDp, dA_total, dN_total, dAdLogDp_std

def plot_number_dist(smps_daily_mean_df, smps_daily_std_df):
    '''
    Description
    ------------
    Creates a plot of scanotron data.

    Parameters
    ------------
    smps_daily_mean_df : pandas dataframe
        Size distribution data where rows are size bins and columns are the mean of daily (or other timespan) data.
    smps_daily_std_df : pandas dataframe
        Size distribution standadr deviation data where rows are size bins and columns are the mean of daily (or other timespan) data.

    Returns
    ------------
    fig
        A plotly graph object.
    '''
    # create melted df
    # create melted df
    smps_daily_mean_melt = pd.melt(smps_daily_mean_df.reset_index(),id_vars=['Dp'], value_name='counts', var_name='sample')
    smps_daily_std_melt = pd.melt(smps_daily_std_df.reset_index(),id_vars=['Dp'], value_name='counts', var_name='sample')
    
    smps_daily_mean_melt['sample'] = smps_daily_mean_melt['sample'].astype(str)
    smps_daily_std_melt['sample'] = smps_daily_std_melt['sample'].astype(str)

    tick_width=1
    tick_loc='inside'
    line_color='black'
    line_width=1

    xticks=7
    x_tick_angle=45
    x_title_size=14
    x_tick_font_size=12
    x_tick_font_color='black'
    x_title_color='black'

    yticks=7
    y_title_size=16
    y_tick_font_size=14
    y_title_color="black"
    y_tick_font_color = "black"

    fig=px.line(smps_daily_mean_melt, x='Dp', y='counts', color='sample', facet_col='sample',
     facet_col_wrap=5, error_y=smps_daily_std_melt['counts'])

    fig.update_xaxes(type='log', title='D<sub>p</sub> (nm)')
    fig.update_yaxes(type='log', title='dN/dLogD<sub>p</sub> (particles/cm<sup>3</sup>)')

    fig.update_yaxes(range=[1,3.8],titlefont=dict(size=y_title_size, color=y_title_color), 
                    ticks=tick_loc, nticks=yticks, tickwidth=tick_width, 
                    tickfont=dict(size=y_tick_font_size, color=y_tick_font_color),
                    showline=True, linecolor=line_color, linewidth=line_width)

    fig.update_xaxes(showticklabels=True,ticks=tick_loc, nticks=xticks, tickwidth=tick_width, showline=True, 
                    linecolor=line_color,linewidth=line_width, 
                    titlefont=dict(size=x_title_size, color=x_title_color), 
                    tickfont=dict(size=x_tick_font_size, color=x_tick_font_color))


    #for anno in fig['layout']['annotations']:
    #    anno['text']=anno.text.split('=')[1]
        
    #for axis in fig.layout:
     #   if type(fig.layout[axis]) == go.layout.YAxis and fig.layout[axis].anchor not in ['x','x16', 'x21', 'x11', 'x6']:
      #      fig.layout[axis].title.text = ''

            
    fig.update_traces(mode='markers+lines')

    fig.update_layout(showlegend=False, width=1100, height=1300, template='plotly_white')
    
    return fig
    #fig.show(renderer="jpg")
    #fig.write_image("manuscripts\\IN\\FIGURES\\figS4.png")

def plot_surface_dist(dAdLogDp, dAdLogDp_std):
    dAdLogDpMelt=dAdLogDp.reset_index().melt(id_vars='Dp')
    dAdLogDpMelt['variable']=dAdLogDpMelt['variable'].astype(str)

    dAdLogDpMelt_std = dAdLogDp_std.reset_index().melt(id_vars='Dp')
    dAdLogDpMelt_std['variable'] = dAdLogDpMelt_std['variable'].astype(str)

    dAdLogDpMelt['value_2'] = dAdLogDpMelt['value'] * 1e-6
    dAdLogDpMelt_std['value_2'] = dAdLogDpMelt_std['value'] * 1e-6

    tick_width=1
    tick_loc='inside'
    line_color='black'
    line_width=1

    xticks=7
    x_tick_angle=45
    x_title_size=14
    x_tick_font_size=12
    x_tick_font_color='black'
    x_title_color='black'

    yticks=7
    y_title_size=16
    y_tick_font_size=14
    y_title_color="black"
    y_tick_font_color = "black"

    fig=px.line(dAdLogDpMelt, x='Dp', y='value_2', color='variable',facet_col='variable', facet_col_wrap=5,
            error_y=dAdLogDpMelt_std['value_2'])
    fig.update_xaxes(type='log', title='Dp (nm)')
    fig.update_yaxes(title='dA/dLogDp (\u03BCm<sup>2</sup>/cm<sup>3</sup>)', type='log')

    fig.update_yaxes(range=[-1,3],titlefont=dict(size=y_title_size, color=y_title_color), 
                    ticks=tick_loc, nticks=yticks, tickwidth=tick_width, 
                    tickfont=dict(size=y_tick_font_size, color=y_tick_font_color),
                    showline=True, linecolor=line_color, linewidth=line_width)

    fig.update_xaxes(showticklabels=True,ticks=tick_loc, nticks=xticks, tickwidth=tick_width, showline=True, 
                    linecolor=line_color,linewidth=line_width, 
                    titlefont=dict(size=x_title_size, color=x_title_color), 
                    tickfont=dict(size=x_tick_font_size, color=x_tick_font_color))

    for anno in fig['layout']['annotations']:
        anno['text']=anno.text.split('=')[1]
        
    for axis in fig.layout:
        if type(fig.layout[axis]) == go.layout.YAxis and fig.layout[axis].anchor not in ['x','x16', 'x21', 'x11', 'x6']:
            fig.layout[axis].title.text = ''

    fig.update_traces(mode='markers+lines')
    fig.update_layout(showlegend=False, width=1100, height=1300, template='plotly_white')
    return fig
    #fig.write_image("manuscripts\\IN\\FIGURES\\response_figs\\figS4.png", scale=4)

def wilsonLower(p, n=26, z = 1.96):
    '''
    p is the frozen fraction
    n is number of tubes
    z is confidence level
    '''
    try:
        denominator = 1 + z**2/n
        centre_adjusted_probability = p + z*z / (2*n)
        adjusted_standard_deviation = math.sqrt((p*(1 - p) + z*z / (4*n)) / n)

        lower_bound = (centre_adjusted_probability - z*adjusted_standard_deviation) / denominator
        upper_bound = (centre_adjusted_probability + z*adjusted_standard_deviation) / denominator

        return(lower_bound)
    except ValueError:
        return(0)

def wilsonUpper(p, n=26, z = 1.96):
    try:
        denominator = 1 + z**2/n
        centre_adjusted_probability = p + z*z / (2*n)
        adjusted_standard_deviation = math.sqrt((p*(1 - p) + z*z / (4*n)) / n)

        lower_bound = (centre_adjusted_probability - z*adjusted_standard_deviation) / denominator
        upper_bound = (centre_adjusted_probability + z*adjusted_standard_deviation) / denominator

        return(upper_bound)
    except ValueError:
        return(0)

def calculate_wilson_errors(project, location, type_, n = 26):
    '''
    Description
    ------------
    Takes calculated report files and creates a csv of error bars. The csv is saved in the same location as the cleaned combined time series data file.
    lower and upper bounds for blank subtracted frozen fraction of tubes are calculated using subfunctions (upperBound, lowerBound, respectively). 
    These fractions are then converted to a number of blank subtracted tubes that are frozen (upper_N-BLNK, lower_N-BLNK, respectively).
    These bounds are then converted into INP/tube upper and lower bounds. Then they are converted to IN/mL and IN/L upper and lower bounds.
    Finally, the difference between each bound and the original observed value is calculated to determine the size of the error bars.

    Paths
    ------------
    raw input data: \\[PROJECT_ROOT]\\data\\interim\\IN\\calculated\\[SAMPLE_TYPE]\\[SAMPLE_LOCATION]\\[FILE]
    cleaned output file: \\[PROJECT_ROOT]\\data\\interim\\IN\\cleaned\\combinedtimeseries\\[SAMPLE_TYPE]\\[FILE]

    Recent Updates
    -------------
    02/11/2020 - Added documentation.
    
    Parameters
    ------------
        project : str
            The project. This needs to be defined since projects before Sea2Cloud were saved in different formats. [me3, nz2020]
        location : str
            Where sample was collected. [uway, ASIT, wkbtsml, wkbtssw, bubbler, coriolis]
        type_ : str
            Description
        n : int
            Number of tubes.
    '''
    error_all = pd.DataFrame()
  

    # cell_vol --> given in ml
    #  In here, calculate the error_y and minus_y. Send both the bound and error_y, minus_y values all together so they can be used however.
    if project == 'me3':
        for proc in ['FILTERED','UNFILTERED']:
            for file in os.listdir("..\\data\\interim\\IN\\calculated\\"+type_+"\\"+proc+'\\'):
                if file.endswith('.xls'):
                
                    error=pd.DataFrame()
                    singleFile= pd.read_excel('..\\data\\interim\\IN\\calculated\\'+type_+'\\'+proc+'\\'+file, sheet_name='summary', header=5)
                    singleFile=singleFile.loc[:,'T (*C)':]    

                    singleFile['lowerBound']=singleFile['FrozenFraction'].apply(wilsonLower)
                    singleFile['upperBound']=singleFile['FrozenFraction'].apply(wilsonUpper)
                    
                    singleFile['upper_N-BLNK']=singleFile['upperBound']*n
                    singleFile['lower_N-BLNK']=singleFile['lowerBound']*n

                    singleFile['IN/tube_upper']=numpy.log(n)-numpy.log(n-singleFile['upper_N-BLNK'])
                    singleFile['IN/tube_lower']=numpy.log(n)-numpy.log(n-singleFile['lower_N-BLNK'])
                    
                    singleFile['IN/ml_upper']=singleFile['IN/tube_upper']/.2
                    singleFile['IN/ml_lower']=singleFile['IN/tube_lower']/.2
                    
                    singleFile['IN/L_upper']=singleFile['IN/ml_upper']*1000
                    singleFile['IN/L_lower']=singleFile['IN/ml_lower']*1000

                    error['IN/L_lower']=singleFile['IN/L_lower']
                    error['IN/L_upper']=singleFile['IN/L_upper']
                
                    error['error_y'] = singleFile['IN/L_upper'] - singleFile['IN/L']
                    error['error_minus_y'] = abs(singleFile['IN/L_lower'] - singleFile['IN/L'])

                    error.index = singleFile['T (*C)'].round(1)

                    error = error.T
                    error.index.rename('value_name', inplace=True)
                    error.reset_index(inplace=True)
                    name = file.split('-')
                    
                    bag = name[2][1]
                    day = name[3]
                    day = day.replace('D','')
                    day = day.replace('.xls','')
                    day = int(day)
                    
                    error['day'] = day
                    error['bag'] = bag
                    error['filtered/unfiltered'] = proc

                    error_all = pd.concat([error_all, error])

        error_all.set_index(['day','bag','value_name','filtered/unfiltered'],inplace=True)        
        error_all.to_csv('C:\\Users\\trueblood\\projects\\me3\\data\\interim\\IN\\cleaned\\seawater\\IN_error_wilson.csv')


    else:
        for file in os.listdir("..\\data\\interim\\IN\\calculated\\"+type_+"\\"+location+'\\'):
            if file.endswith('.xlsx'):
                for proc in ['UH','H']:
                    if type_ == 'seawater':
                        error=pd.DataFrame()
                        singleFile= pd.read_excel('..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'+file, sheet_name='summary_UF_'+proc, header=0)   

                        singleFile['lowerBound']=singleFile['FrozenFraction'].apply(wilsonLower)
                        singleFile['upperBound']=singleFile['FrozenFraction'].apply(wilsonUpper)
                        
                        singleFile['upper_N-BLNK']=singleFile['upperBound']*n
                        singleFile['lower_N-BLNK']=singleFile['lowerBound']*n

                        singleFile['IN/tube_upper']=numpy.log(n)-numpy.log(n-singleFile['upper_N-BLNK'])
                        singleFile['IN/tube_lower']=numpy.log(n)-numpy.log(n-singleFile['lower_N-BLNK'])
                        
                        singleFile['IN/ml_upper']=singleFile['IN/tube_upper']/.2
                        singleFile['IN/ml_lower']=singleFile['IN/tube_lower']/.2
                        
                        singleFile['IN/L_upper']=singleFile['IN/ml_upper']*1000
                        singleFile['IN/L_lower']=singleFile['IN/ml_lower']*1000

                        error['IN/L_lower']=singleFile['IN/L_lower']
                        error['IN/L_upper']=singleFile['IN/L_upper']
                    
                        error['error_y'] = singleFile['IN/L_upper'] - singleFile['IN/L']
                        error['error_minus_y'] = abs(singleFile['IN/L_lower'] - singleFile['IN/L'])

                        error.index = singleFile['T (*C)'].round(1)

                        error = error.T
                        error.index.rename('value_name', inplace=True)
                        error.reset_index(inplace=True)
                        name = file.split('_')
                    
                        error['type'] = name[0]
                        error['location']=name[1]
                        error['filtered']=name[2]
                        error['date']=name[3]
                        error['time']=name[4][0:2]+'h'+name[4][2:4]
                        error['datetime_str'] = error.date.astype(str)+' '+error.time
                        error['datetime'] = pd.to_datetime(error['datetime_str'], format='%d%m%y %Hh%M')
                        error['process'] = proc
                    
                    elif type_ == 'aerosol':
                        error=pd.DataFrame()
                        singleFile= pd.read_excel('..\\data\\interim\\IN\\calculated\\'+type_+'\\'+location+'\\'+file, sheet_name='summary_UF_'+proc, header=0)   

                        error['IN/L_lower']=singleFile['lower INP/L']
                        error['IN/L_upper']=singleFile['upper INP/L']
                    
                        error['error_y'] = error['IN/L_upper'] - singleFile['IN/L']
                        error['error_minus_y'] = abs(error['IN/L_lower'] - singleFile['IN/L'])

                        error.index = singleFile['T (*C)'].round(1)

                        error = error.T
                        error.index.rename('value_name', inplace=True)
                        error.reset_index(inplace=True)
                        name = file.split('_')
                        error['type'] = name[0]
                        error['location']=name[1]
                        error['filtered']=name[2]
                        error['size'] = name[3]
                        error['date']=name[4]
                        error['time']=name[5][0:2]+'h'+name[5][2:4]
                        error['datetime_str'] = error.date.astype(str)+' '+error.time
                        error['datetime'] = pd.to_datetime(error['datetime_str'], format='%d%m%y %Hh%M')
                        error['process'] = proc
                    error_all = pd.concat([error_all, error])
        
        strt=error_all['date'].min()[0:8]
        end=error_all['date'].max()[0:8]
        out_name = strt+'_'+end
        error_all = error_all.drop(columns='datetime_str')
        error_all = error_all.drop(columns='date')
        error_all.set_index(['type','location','filtered','value_name','time','process'],inplace=True)        
        error_all.to_csv('..\\data\\interim\\IN\\cleaned\\combinedtimeseries\\'+type_+'\\'+location+'_'+out_name+'wilson_error.csv')

def load_wilson_errors():
    '''
    See IN_Analysis_V1.ipynb. This will load and clean up the spreadsheets of error bars.
    '''
    pd.read_csv("C:\\Users\\trueblood\\projects\\me3\\data\\interim\\IN\\cleaned\\seawater\\IN_error_wilson.csv", sep=',', index_col=[0,1,2])
    errors_melt = pd.melt(errors.reset_index(), id_vars=['bag','day', 'value_name'], value_name='error', var_name='T')

def plot_sml_inp(inp_df):
    '''
    This accepts a dataframe of INP values along with their uncertainties and plots heated vs unheated as well as literature values.
    '''
    tick_width=1
    tick_loc='inside'
    line_color='black'
    line_width=1

    xticks=10
    x_tick_angle=45
    x_title_size=15
    x_tick_font_size=14
    x_tick_font_color='black'
    x_title_color='black'

    y_title_size=15
    y_tick_font_size=10
    y_title_color="black"
    y_tick_font_color = "black"

    fig = go.Figure()


    fig.add_trace(go.Scatter(mode='markers',
            x = inp_df[inp_df['process']=='UH']['temp'], y = inp_df[inp_df['process']=='UH']['inp/l'], name = 'Unheated SML', 
            error_y=dict(
                    type='data',
                    symmetric=False,
                    array=inp_df[inp_df['process']=='UH']['error_y'],
                    arrayminus=inp_df[inp_df['process']=='UH']['error_minus_y'],
                    thickness=1.5,
                    width=5)
            ))

    fig.add_trace(go.Scatter(mode='markers',
        x = inp_df[inp_df['process']=='H']['temp'], y = inp_df[inp_df['process']=='H']['inp/l'], name = 'Heated SML', 
        error_y=dict(
                type='data',
                symmetric=False,
                array=inp_df[inp_df['process']=='H']['error_y'],
                arrayminus=inp_df[inp_df['process']=='H']['error_minus_y'],
                thickness=1.5,
                width=5)
        ))


    fig.add_trace(go.Scatter(mode='lines',x = [-15,-20,-25,-22, -17, -15, -15], y = [0.6*10**3,0.6*10**3,60*10**3,60*10**3, 2*10**3, 2*10**3, 0.6*10**3], name = 'McCluskey et al. (2018) (Southern Ocean)'))

    fig.add_trace(go.Scatter(mode='lines',x = [-5,-16,-25,-10, -5], y = [3*10**4,3*10**4,4*10**6,4*10**6, 3*10**4], name = 'Wilson et al. (2015)'))

    fig.add_trace(go.Scatter(mode='lines', x = [-10,-22,-28,-22, -15, -10], y = [3*10**4,3*10**4,1*10**7,1*10**7, 1*10**6, 3*10**4], name = 'Irish et al. (2017)'))

    fig.add_trace(go.Scatter(mode='lines', x = [-5,-13,-16,-17, -11, -5], y = [2*10**4,2*10**4,2*10**5,4*10**6, 4*10**6, 2*10**4], name = 'Irish et al. (2019)'))

    fig.add_trace(go.Scatter(mode='lines', x = [-9,-15,-16,-27, -24, -9], y = [1.8*10**2,1.8*10**2,2*10**3,4*10**6, 4*10**6, 1.8*10**2], name = 'Gong et al. (2020)'))

    fig.add_trace(go.Scatter(mode='lines', x = [-10,-15,-16.5,-16.5, -15, -10], y = [1.9*10**2,1.9*10**2,9*10**2,2*10**4, 2*10**4, 1.9*10**2], name = 'Trueblood et al. (2020) - Microlayer'))
    
    fig.update_yaxes(exponentformat='power', type='log')

    fig.update_yaxes(nticks=20, range=[1.3,5], title="INP/L",
                    titlefont=dict(size=y_title_size, color=y_title_color), 
                    ticks=tick_loc, tickwidth=tick_width, tickfont=dict(size=y_tick_font_size, color=y_tick_font_color),
                    showline=True, linecolor=line_color, linewidth=line_width)


    fig.update_xaxes(title='Temperature (\u00B0C)', range=[-20,-3], ticks=tick_loc, nticks=xticks, tickwidth=tick_width,
                    showline=True, linecolor=line_color, linewidth=line_width,
                    tickangle=x_tick_angle, tickfont=dict(size=x_tick_font_size, color=x_tick_font_color),
                    title_font=dict(size=x_title_size, color=x_title_color),
                    )

    fig.update_traces(marker=dict(size=5,line = dict(width=.5, color='black')
                    ))

    fig.update_layout(template='plotly_white', height=600, width=700, showlegend=True,  
            legend=dict(title='',font=dict(size=14, color='black'), x=0.62, y=.5, bordercolor="Black", borderwidth=1))
    
    return fig

def plot_ssw_inp(inp_df):
    '''
    This accepts a dataframe of INP values along with their uncertainties and plots heated vs unheated as well as literature values.
    '''
    tick_width=1
    tick_loc='inside'
    line_color='black'
    line_width=1

    xticks=10
    x_tick_angle=45
    x_title_size=15
    x_tick_font_size=14
    x_tick_font_color='black'
    x_title_color='black'

    y_title_size=15
    y_tick_font_size=10
    y_title_color="black"
    y_tick_font_color = "black"

    fig = go.Figure()


    fig.add_trace(go.Scatter(mode='markers',
            x = inp_df[inp_df['process']=='UH']['temp'], y = inp_df[inp_df['process']=='UH']['inp/l'], name = 'Unheated SSW', 
            error_y=dict(
                    type='data',
                    symmetric=False,
                    array=inp_df[inp_df['process']=='UH']['error_y'],
                    arrayminus=inp_df[inp_df['process']=='UH']['error_minus_y'],
                    thickness=1.5,
                    width=5)
            ))

    fig.add_trace(go.Scatter(mode='markers',
        x = inp_df[inp_df['process']=='H']['temp'], y = inp_df[inp_df['process']=='H']['inp/l'], name = 'Heated SSW', 
        error_y=dict(
                type='data',
                symmetric=False,
                array=inp_df[inp_df['process']=='H']['error_y'],
                arrayminus=inp_df[inp_df['process']=='H']['error_minus_y'],
                thickness=1.5,
                width=5)
        ))


    fig.add_trace(go.Scatter(mode='lines',x = [-15,-20,-25,-22, -17, -15, -15], y = [0.6*10**3,0.6*10**3,60*10**3,60*10**3, 2*10**3, 2*10**3, 0.6*10**3], name = 'McCluskey et al. (2018) (Southern Ocean)'))

    fig.add_trace(go.Scatter(mode='lines',x = [-5,-16,-25,-10, -5], y = [3*10**4,3*10**4,4*10**6,4*10**6, 3*10**4], name = 'Wilson et al. (2015)'))

    fig.add_trace(go.Scatter(mode='lines', x = [-10,-22,-28,-22, -15, -10], y = [3*10**4,3*10**4,1*10**7,1*10**7, 1*10**6, 3*10**4], name = 'Irish et al. (2017)'))

    fig.add_trace(go.Scatter(mode='lines', x = [-5,-13,-16,-17, -11, -5], y = [2*10**4,2*10**4,2*10**5,4*10**6, 4*10**6, 2*10**4], name = 'Irish et al. (2019)'))

    fig.add_trace(go.Scatter(mode='lines', x = [-9,-15,-16,-27, -24, -9], y = [1.8*10**2,1.8*10**2,2*10**3,4*10**6, 4*10**6, 1.8*10**2], name = 'Gong et al. (2020)'))
    
    fig.add_trace(go.Scatter(mode='lines', x = [-13,-15,-16.5,-16.5, -14, -13], y = [1.9*10**2,1.9*10**2,4*10**2,4*10**3, 2*10**3, 1.9*10**2], name = 'Trueblood et al. (2020) - Seawater'))

    fig.update_yaxes(exponentformat='power', type='log')

    fig.update_yaxes(nticks=20, range=[1.3,5], title="INP/L",
                    titlefont=dict(size=y_title_size, color=y_title_color), 
                    ticks=tick_loc, tickwidth=tick_width, tickfont=dict(size=y_tick_font_size, color=y_tick_font_color),
                    showline=True, linecolor=line_color, linewidth=line_width)


    fig.update_xaxes(title='Temperature (\u00B0C)', range=[-20,-3], ticks=tick_loc, nticks=xticks, tickwidth=tick_width,
                    showline=True, linecolor=line_color, linewidth=line_width,
                    tickangle=x_tick_angle, tickfont=dict(size=x_tick_font_size, color=x_tick_font_color),
                    title_font=dict(size=x_title_size, color=x_title_color),
                    )

    fig.update_traces(marker=dict(size=5,line = dict(width=.5, color='black')
                    ))

    fig.update_layout(template='plotly_white', height=600, width=700, showlegend=True,  
            legend=dict(title='',font=dict(size=14, color='black'), x=0.62, y=.5, bordercolor="Black", borderwidth=1))
    
    return fig

def CleanAqualog(instr, outpath):
    '''
    Summary:
    
    Loads all raw aqualog data files for a given instrument (aqlog1 or aqlog2) and cleans it up.
    Returns the cleaned dataset to chosen outpath. 
    
    The steps of the cleaning process are as follows:
        1) open all data files in te data/raw folder path
        2) append all data files into one df
        3) remove bad chars in column names
        4) create a timeString column in UTC time
        5) save df to csv in specified folder
    
    Parameters:
    
    instr (string): aqualog1 or aqualog2
    outpath (string): location where cleaned csv file is saved.
    
    Returns:
    
    dfBig (df): the df that was just saved to a folder
    outName (str): string of the start and end datetimes
    
    '''
    dfBig=pd.DataFrame()
    path='C:\\Users\\trueblood\\projects\\nz2020\\New-Zealand-2020\\data\\raw\\'+instr+'\\'

    #Read in all the files in a given folder.
    for file in os.listdir(path):
        if file.endswith('.csv'):
            df = pd.read_csv(path+file, sep='\t', parse_dates=['# UTC ISO8601'])
            dfBig = dfBig.append(df)
            # Read in column names
            columns = pd.read_csv(path+file,nrows=0,sep='\t').columns.tolist()
            # Remove all bad chars from column names
            for name in range(len(columns)):
                columns[name]=(columns[name]).strip()
                columns[name]=(columns[name]).replace("#","")
                columns[name]=(columns[name]).replace(" ","")
                columns[name]=columns[name].lower()
                columns[name]=(columns[name]).replace("utciso8601","time")
    dfBig.columns = columns
    dfBig['timeString'] = dfBig['time'].dt.strftime('%Y-%m-%d %H:%M:%S')
    dfBig=dfBig.set_index(['timeString'])
    dfBig.time=dfBig.time.dt.tz_localize(None)
    strt=dfBig.index[0]
    end=dfBig.index[-1]
    outName = strt[0:10]+'_'+end[0:10]
    dfBig.to_csv(outpath+outName+'.csv')
    return dfBig, outName